import os
import logging
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from dotenv import load_dotenv
import openai
import re
import json
import typing
import uuid

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

openai.api_key = os.getenv("OPENAI_API_KEY")


from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from io import BytesIO
from fastapi import UploadFile, File, Form
from typing import List

app = FastAPI()

# Allow CORS from frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5174"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Request model
class HSCRequest(BaseModel):
    description: str
    country: str = "US"

# def format_trade_code(code: str) -> str:
# Expected code lengths by country
COUNTRY_CODE_LENGTHS = {
    "US": 10,
    "EU": 10,
    "GB": 10,
    "CA": 10,
    "CN": 10,
    "JP": 9,
    "IN": 8,
    "AU": 8,
    "MX": 8,
    "BR": 8,
}

# Helper to format raw code string into standardized dotted format
def format_trade_code(code: str, country: str = None) -> str:
    # Remove any non-digit characters
    digits = re.sub(r"\D", "", code)
    length = len(digits)
    # Enforce expected length if country specified
    if country:
        c = country.upper()
        exp = COUNTRY_CODE_LENGTHS.get(c)
        if exp:
            # Truncate or pad to expected length
            if length > exp:
                digits = digits[:exp]
            elif length < exp:
                digits = digits.ljust(exp, '0')
            length = exp
    # Determine grouping based on length
    parts: typing.List[str] = []
    if length == 6:
        parts = [digits[0:2], digits[2:4], digits[4:6]]
    elif length == 8:
        parts = [digits[0:4], digits[4:6], digits[6:8]]
    elif length == 10:
        parts = [digits[0:4], digits[4:6], digits[6:8], digits[8:10]]
    else:
        parts = [digits[i:i+2] for i in range(0, length, 2) if digits[i:i+2]]
    return ".".join(parts)

# Helper function to get country code length
def get_country_code_length(country: str) -> int:
    country = country.upper()
    return COUNTRY_CODE_LENGTHS.get(country)

# Helper function to create system message
def create_system_message(country: str, exp: int = None) -> str:
    country = country.upper()
    if exp:
        instr = f"Using the official {country} tariff schedule, provide the valid {exp}-digit trade classification code"
    else:
        instr = f"Using the official {country} tariff schedule, provide the valid trade classification code"
    return (
        "You are an expert in international trade classifications. "
        f"{instr} for the following product. "
        "Respond with only the numeric code, no additional text or explanation."
    )

# Function to get HSC code from OpenAI
def get_hsc_from_openai(description: str, country: str = "US") -> str:
    exp = get_country_code_length(country)
    system_message = create_system_message(country, exp)
    user_prompt = f"Product description: '{description}'"

    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_prompt}
            ],
            max_tokens=10,
            temperature=0,
            stop=["\n"]
        )
        raw = response.choices[0].message.content
        # Extract a 6-, 8-, or 10-digit code if present
        match = re.search(r"\b(\d{6}(?:\d{2}){0,2})\b", raw)
        code_raw = match.group(1) if match else raw.strip()
        return format_trade_code(code_raw, country)
    except Exception as e:
        logger.error(f"OpenAI API error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve HSC code from OpenAI.")
    
# Function to get detailed HSC info from OpenAI as structured JSON
def get_hsc_details(description: str, country: str = "US") -> dict:
    exp = get_country_code_length(country)
    # Build system instruction with expected digit count for code key
    if exp:
        code_desc = f"code (the {exp}-digit classification code as a string)"
    else:
        code_desc = "code (the classification code as a string)"
    system_message = (
        "You are an expert in international trade classifications. "
        f"Using the official {country} tariff schedule, provide a JSON object with the following keys: "
        f"{code_desc}, "
        "title (a concise classification title), "
        "description (the official classification description), "
        "category (a broad product category), and "
        "match (an integer from 0 to 100 indicating confidence). "
        "Respond with only the JSON object, no additional text or formatting."
    )
    user_prompt = f"Product description: '{description}'"
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0,
            max_tokens=200
        )
        content = response.choices[0].message.content.strip()
        # Remove code fences if present
        if content.startswith("```"):
            parts = content.split("```")
            if len(parts) >= 3:
                content = parts[1].strip()
        # Parse JSON and format the code field
        data = json.loads(content)
        if "code" in data:
            data["code"] = format_trade_code(data["code"], country)
        return data
    except Exception as e:
        logger.error(f"Error parsing HSC details JSON: {e}")
        # Fallback: minimal response
        code = get_hsc_from_openai(description, country)
        return {
            "code": code,
            "title": "",
            "description": description,
            "category": "",
            "match": 0
        }

@app.post("/get-hsc")
def get_hsc(request: HSCRequest):
    description = request.description.strip()
    country = request.country.strip()
    if not description:
        logger.warning("Empty description received.")
        raise HTTPException(status_code=400, detail="Description is required.")
    # Retrieve structured HSC/HTS details based on country
    details = get_hsc_details(description, country)
    # Attach a unique ID
    return {"id": str(uuid.uuid4()), **details}

@app.get("/health")
async def health_check():
    return {"status": "OK"}


# Data model for QA requests
class TableRow(BaseModel):
    description: str
    hsc_code: str

class QARequest(BaseModel):
    question: str
    table: List[TableRow]

@app.post("/ask")
def ask(request: QARequest):
    # Build a text representation of the table
    lines = [f"{i+1}. {row.description} -> {row.hsc_code}" for i, row in enumerate(request.table)]
    table_text = "\n".join(lines)
    system_msg = (
        "You are a helpful assistant that answers questions based on a table of product descriptions and HSC codes. "
        "Use only the information provided; do not invent data."
    )
    user_prompt = f"Here is the table of products and HSC codes:\n{table_text}\n\nQuestion: {request.question}"
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0
        )
        answer = response.choices[0].message.content.strip()
        return {"answer": answer}
    except Exception as e:
        logger.error(f"OpenAI QA API error: {e}")
        raise HTTPException(status_code=500, detail="Failed to answer question.")

@app.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    country: str = Form("US")
):
    # Accept only Excel files
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx files are supported.")
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents))
        ws = wb.active
        # Read header row to find 'description' column
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        desc_idx = next((i for i, h in enumerate(header) if h and str(h).lower() == 'description'), 0)
        # Process each row and build result list
        results = []  # type: List[dict]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            desc_cell = row[desc_idx].value
            desc = str(desc_cell) if desc_cell is not None else ''
            # Classify description per selected country
            code = get_hsc_from_openai(desc, country) if desc else ''
            # Stub confidence; refine as needed
            confidence = 1.0 if code else 0.0
            results.append({"description": desc, "hsc_code": code, "confidence": confidence})
        return {"filename": file.filename, "rows": results}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing Excel upload: {e}")
        raise HTTPException(status_code=500, detail="Failed to process Excel file.")

@app.post("/bulk-hsc")
async def bulk_hsc_generation(
    descriptions: List[str],
    country: str = "US"
):
    """Generate HSC codes for multiple descriptions in batch"""
    try:
        results = []
        for description in descriptions:
            if description and description.strip():
                try:
                    details = get_hsc_details(description.strip(), country)
                    results.append({
                        "description": description.strip(),
                        "hsc_code": details["code"],
                        "confidence": details["match"],
                        "category": details["category"],
                        "title": details["title"]
                    })
                except Exception as e:
                    logger.error(f"Error processing description '{description}': {e}")
                    results.append({
                        "description": description.strip(),
                        "hsc_code": "Error",
                        "confidence": 0,
                        "category": "",
                        "title": ""
                    })
            else:
                results.append({
                    "description": description,
                    "hsc_code": "",
                    "confidence": 0,
                    "category": "",
                    "title": ""
                })
        
        return {
            "results": results,
            "total_processed": len(results),
            "successful": len([r for r in results if r["hsc_code"] and r["hsc_code"] != "Error"])
        }
    except Exception as e:
        logger.error(f"Error in bulk HSC generation: {e}")
        raise HTTPException(status_code=500, detail="Failed to process bulk HSC generation.")